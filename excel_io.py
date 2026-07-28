"""
Lectura/escritura del Excel preservando el original y anadiendo las dos
columnas de salida: CONCILIADO (booleano) e INFORMAÇÃO (referencia).
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import reconciliation as rec


COL_CONCILIADO = "CONCILIADO"
COL_INFO = "INFORMAÇÃO"


def _find_header_row(ws, max_scan: int = 15):
    """Encuentra la fila de cabeceras buscando 'DEB_CRED' / 'DATA_CERTA'."""
    for r in range(1, min(max_scan, ws.max_row) + 1):
        headers = [c.value for c in ws[r]]
        cols = rec.detectar_columnas(headers)
        if cols["valor"] and cols["fecha"]:
            return r, headers, cols
    # fallback: primera fila
    headers = [c.value for c in ws[1]]
    return 1, headers, rec.detectar_columnas(headers)


def procesar_workbook(data: bytes, cfg: rec.ReconConfig | None = None,
                      hoja: str | None = None,
                      override_valor: str | None = None,
                      override_fecha: str | None = None):
    """
    Procesa el .xlsx (en bytes) y devuelve:
        (bytes_resultado, resultado_rec, info)
    info incluye la hoja usada y las columnas detectadas.
    """
    cfg = cfg or rec.ReconConfig()
    wb = load_workbook(io.BytesIO(data))
    ws = wb[hoja] if hoja else wb.active

    header_row, headers, cols = _find_header_row(ws)
    col_valor = override_valor or cols["valor"]
    col_fecha = override_fecha or cols["fecha"]

    if not col_valor:
        raise ValueError(
            "No se encontro la columna de importes 'DEB_CRED'. "
            f"Cabeceras encontradas: {headers}"
        )

    # indices de columna (1-based) a partir de la cabecera
    header_map = {}
    for idx, h in enumerate(headers, start=1):
        if h is not None:
            header_map[str(h)] = idx
    ci_valor = header_map.get(str(col_valor))
    ci_fecha = header_map.get(str(col_fecha)) if col_fecha else None

    # leer filas de datos
    filas = []
    for r in range(header_row + 1, ws.max_row + 1):
        valor_cell = ws.cell(row=r, column=ci_valor).value
        fecha_cell = ws.cell(row=r, column=ci_fecha).value if ci_fecha else None
        # ignorar filas totalmente vacias
        if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, len(headers) + 1)):
            continue
        filas.append({
            "row": r,
            "valor": rec._to_cents(valor_cell),
            "fecha": pd.to_datetime(fecha_cell, errors="coerce", dayfirst=True)
            if fecha_cell is not None else None,
        })

    resultado = rec.conciliar(filas, cfg)
    resultado.col_valor = col_valor
    resultado.col_fecha = col_fecha

    # ---- escribir columnas de salida ----
    last_col = len(headers)
    # reutilizar columnas si ya existen, si no anadir al final
    col_conc = header_map.get(COL_CONCILIADO)
    col_inf = header_map.get(COL_INFO)
    if col_conc is None:
        last_col += 1
        col_conc = last_col
    if col_inf is None:
        last_col += 1
        col_inf = last_col

    head_fill = PatternFill("solid", fgColor="D9E1F2")
    head_font = Font(bold=True)
    ok_fill = PatternFill("solid", fgColor="E2EFDA")     # verde suave
    no_fill = PatternFill("solid", fgColor="FCE4D6")     # naranja suave

    for c, name in ((col_conc, COL_CONCILIADO), (col_inf, COL_INFO)):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center")

    for f in filas:
        a = resultado.asignacion[f["row"]]
        cc = ws.cell(row=f["row"], column=col_conc, value=bool(a["conciliado"]))
        cc.alignment = Alignment(horizontal="center")
        ci = ws.cell(row=f["row"], column=col_inf, value=a["referencia"])
        fill = ok_fill if a["conciliado"] else no_fill
        cc.fill = fill
        ci.fill = fill

    ws.column_dimensions[get_column_letter(col_conc)].width = 13
    ws.column_dimensions[get_column_letter(col_inf)].width = 18

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    info = {
        "hoja": ws.title,
        "fila_cabecera": header_row,
        "col_valor": col_valor,
        "col_fecha": col_fecha,
        "hojas": wb.sheetnames,
    }
    return out.getvalue(), resultado, info


def tabla_detalle(data: bytes, resultado, info) -> pd.DataFrame:
    """DataFrame legible para mostrar en pantalla (todas las columnas + salida)."""
    df = pd.read_excel(io.BytesIO(data), sheet_name=info["hoja"],
                       header=info["fila_cabecera"] - 1)
    return df
