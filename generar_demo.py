"""Genera un Excel de demostracion parecido al caso real del usuario."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# columnas en orden ALTERADO a proposito (para probar deteccion por nombre)
filas = [
    # Fornecedor, Data Lanç., DATA_CERTA, Período, Diário, Doc.Interno, Doc.Externo, Descritivo, Descrição, DEBITO, CREDITO, DEB_CRED
    ["221110078 - Inquieto Mobily", "31-07-2024", "31-07-2024", 2024071, 3, 1,  "2024-07-31 00005 750001", "Nota de Lançamento", "09/07 -Inquieto Mobili", 992.63, None, 992.63],
    ["221110078 - Inquieto Mobily", "31-07-2024", "22-07-2024", 2024071, 56, 50, "2024-07-22 00003 730034", "Nota de Lançamento", "FT 2024/221 Inquieto Mobilidade", None, 992.63, -992.63],
    ["221110078 - Inquieto Mobily", "31-08-2024", "31-08-2024", 2024081, 3, 1,  "2024-08-30 00005 850001", "Nota de Lançamento", "06/08 -Inquieto", 992.63, None, 992.63],
    ["221110078 - Inquieto Mobily", "31-08-2024", "30-08-2024", 2024081, 56, 43, "2024-08-30 00003 830025", "Nota de Lançamento", "FT 2024/250 Inquieto Mobilidade", None, 2108.04, -2108.04],
    ["221110078 - Inquieto Mobily", "30-09-2024", "30-09-2024", 2024091, 3, 1,  "2024-09-30 00005 950001", "Nota de Lançamento", "09/09 -Inquieto", 1079.02, None, 1079.02],
    ["221110078 - Inquieto Mobily", "30-09-2024", "30-09-2024", 2024091, 3, 1,  "2024-09-30 00005 950001", "Nota de Lançamento", "09/09 -Inquieto", 1029.02, None, 1029.02],
    # caso extra: nota de credito + factura + pago que netean a cero
    ["221110078 - Inquieto Mobily", "15-10-2024", "10-10-2024", 2024101, 56, 61, "2024-10-10 00003 100061", "Nota de Lançamento", "FT 2024/300", None, 1500.00, -1500.00],
    ["221110078 - Inquieto Mobily", "15-10-2024", "12-10-2024", 2024101, 57, 62, "2024-10-12 00004 100062", "Nota de Lançamento", "NC 2024/030 (nota credito)", 200.00, None, 200.00],
    ["221110078 - Inquieto Mobily", "15-10-2024", "20-10-2024", 2024101, 3, 5,  "2024-10-20 00005 100063", "Nota de Lançamento", "Pagamento FT300", 1300.00, None, 1300.00],
]

cols = ["Fornecedor", "Data Lanç.", "DATA_CERTA", "Período", "Diário",
        "Doc. Interno", "Doc. Externo", "Descritivo", "Descrição",
        "DEBITO", "CREDITO", "DEB_CRED"]

wb = Workbook()
ws = wb.active
ws.title = "Conta Corrente"
ws.append(cols)
for c in ws[1]:
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor="D9D9D9")
    c.alignment = Alignment(horizontal="center")
for row in filas:
    ws.append(row)
wb.save("demo_conta_corrente.xlsx")
print("demo_conta_corrente.xlsx creado con", len(filas), "lineas")
