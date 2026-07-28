"""
App web para conciliar cuentas corrientes de proveedores.

Sube un Excel, localiza por NOMBRE las columnas 'DEB_CRED' (importes con signo)
y 'DATA_CERTA' (fecha), busca grupos de lineas que SUMAN CERO (factura + pago,
factura + varios pagos, notas de credito), y devuelve el mismo Excel con dos
columnas nuevas: CONCILIADO (VERDADEIRO/FALSO) e INFORMAÇÃO (Referência 001...).
"""

import io

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

import reconciliation as rec
import excel_io

st.set_page_config(page_title="Conciliador de Conta Corrente", page_icon="✅", layout="wide")

st.title("✅ Conciliador de Conta Corrente de Fornecedores")
st.caption(
    "Encontra as linhas cujo **DEB_CRED soma zero** (fatura + pagamento, "
    "faturas com vários pagamentos, notas de crédito) e marca as que ficam soltas."
)

with st.sidebar:
    st.header("⚙️ Opções")
    tol = st.selectbox("Tolerância (arredondamento)",
                       ["± 0,01 €", "Exato (0,00)", "± 0,05 €", "± 0,10 €"], index=0)
    tol_map = {"Exato (0,00)": 0, "± 0,01 €": 1, "± 0,05 €": 5, "± 0,10 €": 10}
    max_lineas = st.slider("Máx. linhas por grupo", 2, 12, 8,
                           help="Tamanho máximo de um grupo (1 fatura + N pagamentos, etc.)")
    st.divider()
    st.markdown(
        "**Como funciona**\n\n"
        "1. Procura as colunas `DEB_CRED` e `DATA_CERTA` pelo nome (a ordem pode variar).\n"
        "2. Casa primeiro os pares 1:1 (importe oposto, data mais próxima).\n"
        "3. Depois procura um pagamento que salda várias faturas.\n"
        "4. O que não casa fica **SEM IDENTIFICAR** e forma o saldo em aberto."
    )

ficheiro = st.file_uploader("📄 Carregue o Excel da conta corrente (.xlsx)",
                            type=["xlsx", "xlsm"])

if ficheiro is None:
    st.info("Carregue um ficheiro para começar. Pode testar com o ficheiro de "
            "demonstração incluído no repositório (`demo_conta_corrente.xlsx`).")
    st.stop()

data = ficheiro.read()

# deteccion de hojas y columnas para permitir ajuste manual
wb = load_workbook(io.BytesIO(data), read_only=True)
hojas = wb.sheetnames
col1, col2 = st.columns(2)
hoja = col1.selectbox("Folha", hojas, index=0)

ws = wb[hoja]
hdr_row, headers, cols_auto = excel_io._find_header_row(ws)
headers_str = [str(h) for h in headers if h is not None]

def _idx(name):
    return headers_str.index(str(name)) if name in headers_str else 0

with col2:
    st.write("")  # espaçador
    with st.expander("Ajustar colunas (opcional)"):
        col_valor = st.selectbox("Coluna de importes (DEB_CRED)", headers_str,
                                 index=_idx(cols_auto["valor"]))
        col_fecha = st.selectbox("Coluna de data (DATA_CERTA)", headers_str,
                                 index=_idx(cols_auto["fecha"]) if cols_auto["fecha"] else 0)

col_valor = locals().get("col_valor", cols_auto["valor"])
col_fecha = locals().get("col_fecha", cols_auto["fecha"])

if not col_valor:
    st.error("Não encontrei a coluna de importes `DEB_CRED`. "
             "Abra o expander 'Ajustar colunas' e selecione-a manualmente.")
    st.stop()

st.success(f"Colunas detetadas — importes: **{col_valor}** · data: **{col_fecha}** "
           f"· folha: **{hoja}** (cabeçalho na linha {hdr_row})")

cfg = rec.ReconConfig(tolerancia_centimos=tol_map[tol], max_lineas_por_grupo=max_lineas)

with st.spinner("A conciliar..."):
    out_bytes, resultado, info = excel_io.procesar_workbook(
        data, cfg=cfg, hoja=hoja,
        override_valor=col_valor, override_fecha=col_fecha,
    )

r = resultado.resumen
m1, m2, m3, m4 = st.columns(4)
m1.metric("Linhas totais", r["lineas_total"])
m2.metric("Conciliadas ✅", r["lineas_casadas"])
m3.metric("Soltas ❌", r["lineas_sueltas"])
m4.metric("Saldo em aberto", f"{r['saldo_abierto_eur']:,.2f} €".replace(",", " "))

st.divider()

# tabla de resultados
df = pd.read_excel(io.BytesIO(out_bytes), sheet_name=hoja, header=hdr_row - 1)

def _pinta(row):
    val = row.get(excel_io.COL_CONCILIADO)
    cor = "background-color: #E2EFDA" if val is True else "background-color: #FCE4D6"
    return [cor] * len(row)

tab1, tab2, tab3 = st.tabs(["📋 Tudo", "❌ Soltas (saldo)", "🔗 Grupos"])
with tab1:
    st.dataframe(df.style.apply(_pinta, axis=1), use_container_width=True, height=430)
with tab2:
    soltas = df[df[excel_io.COL_CONCILIADO] == False]
    st.dataframe(soltas, use_container_width=True, height=430)
    st.caption("Estas são as linhas que não casam com nada e compõem o saldo atual.")
with tab3:
    for g in resultado.grupos:
        st.markdown(f"**{g['referencia']}** — {g['tipo']} · linhas Excel {g['rows']}")

st.divider()
nome_out = (ficheiro.name.rsplit(".", 1)[0]) + "_CONCILIADO.xlsx"
st.download_button("⬇️ Descarregar Excel conciliado", data=out_bytes,
                   file_name=nome_out,
                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                   type="primary")
st.caption("A coluna CONCILIADO vem como VERDADEIRO/FALSO. No Excel pode "
           "selecioná-la e usar Inserir → Caixa de verificação para a tornar clicável.")
